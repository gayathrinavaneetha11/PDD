import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def generate_excel_report(results, filepath="test_report.xlsx"):
    """
    Compiles test execution results into a formatted Excel analysis report.
    If openpyxl is not available, falls back to generating a CSV report.
    """
    if not OPENPYXL_AVAILABLE:
        print("[WARN] openpyxl is not installed. Generating CSV report fallback instead.")
        generate_csv_fallback(results, filepath.replace(".xlsx", ".csv"))
        return

    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Styles Definition
    # ----------------------------------------------------
    font_family = "Arial"
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=12, bold=True, color="1B365D")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=10, bold=True)
    regular_font = Font(name=font_family, size=10)
    kpi_val_font = Font(name=font_family, size=20, bold=True, color="1B365D")
    
    # Status Fonts
    pass_font = Font(name=font_family, size=10, bold=True, color="155724")
    fail_font = Font(name=font_family, size=10, bold=True, color="721C24")

    # Fills
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    title_fill = PatternFill(start_color="0D233A", end_color="0D233A", fill_type="solid")
    zebra_fill = PatternFill(start_color="F7F9FB", end_color="F7F9FB", fill_type="solid")
    kpi_fill = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid")
    
    # Status Fills
    pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    # Alignments
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Borders
    thin_side = Side(border_style="thin", color="CCCCCC")
    double_side = Side(border_style="double", color="1B365D")
    thick_bottom = Side(border_style="medium", color="1B365D")
    
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_bottom_double = Border(bottom=double_side, left=thin_side, right=thin_side)
    border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)

    # ----------------------------------------------------
    # SHEET 1: Summary Dashboard
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Summary Dashboard"
    ws1.views.sheetView[0].showGridLines = True

    # 1. Title Block
    ws1.merge_cells("A1:E2")
    title_cell = ws1["A1"]
    title_cell.value = "SmartSafetyTravel - E2E Appium Test Report"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align

    # Apply title fill styling across merged cells
    for r in range(1, 3):
        for c in range(1, 6):
            ws1.cell(row=r, column=c).fill = title_fill

    # Calculate statistics
    total_tests = len(results)
    passed_tests = sum(1 for x in results if x["status"] == "PASS")
    failed_tests = total_tests - passed_tests
    success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0.0
    total_duration = sum(x["duration_sec"] for x in results)

    # 2. KPI Cards
    kpi_labels = ["Total Tests", "Passed", "Failed", "Success Rate", "Total Duration"]
    kpi_values = [total_tests, passed_tests, failed_tests, f"{success_rate:.1f}%", f"{total_duration:.1f}s"]
    
    # Row 4: KPI Labels
    for i, label in enumerate(kpi_labels, start=1):
        cell = ws1.cell(row=4, column=i)
        cell.value = label
        cell.font = bold_font
        cell.fill = kpi_fill
        cell.alignment = center_align
        cell.border = Border(top=thin_side, left=thin_side, right=thin_side)

    # Row 5: KPI Values
    for i, val in enumerate(kpi_values, start=1):
        cell = ws1.cell(row=5, column=i)
        cell.value = val
        cell.font = kpi_val_font
        cell.fill = kpi_fill
        cell.alignment = center_align
        cell.border = Border(bottom=thin_side, left=thin_side, right=thin_side)
        
        # Color specific KPI values
        if i == 2:  # Passed
            cell.font = Font(name=font_family, size=20, bold=True, color="155724")
        elif i == 3: # Failed
            cell.font = Font(name=font_family, size=20, bold=True, color="721C24")

    # 3. Test Cases List Header
    ws1.cell(row=7, column=1).value = "Test Suite Overview"
    ws1.cell(row=7, column=1).font = section_font
    
    headers = ["Test ID", "Test Name", "Status", "Execution Time (s)", "Timestamp"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws1.cell(row=8, column=col_idx)
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_header

    # 4. Populate Summary Data
    for row_idx, r in enumerate(results, start=9):
        ws1.cell(row=row_idx, column=1, value=r["test_id"]).alignment = center_align
        ws1.cell(row=row_idx, column=2, value=r["name"]).alignment = left_align
        
        # Status styling
        status_cell = ws1.cell(row=row_idx, column=3, value=r["status"])
        status_cell.alignment = center_align
        if r["status"] == "PASS":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        else:
            status_cell.fill = fail_fill
            status_cell.font = fail_font
            
        ws1.cell(row=row_idx, column=4, value=r["duration_sec"]).alignment = right_align
        ws1.cell(row=row_idx, column=5, value=r["timestamp"]).alignment = center_align

        # Apply borders and zebra striping
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, 6):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.border = border_all
            cell.font = regular_font
            if is_even and col_idx != 3: # Skip status cell background override
                cell.fill = zebra_fill

    # Auto-adjust column widths for Sheet 1
    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        # We start from row 4 to ignore the merged title row which spans 5 columns
        for cell in col[3:]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # ----------------------------------------------------
    # SHEET 2: Detailed Execution Logs
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Detailed Execution Logs")
    ws2.views.sheetView[0].showGridLines = True

    # Section Title
    ws2.cell(row=1, column=1).value = "Detailed E2E Test Execution Logs"
    ws2.cell(row=1, column=1).font = section_font

    headers2 = ["Test ID", "Test Case Name", "Description", "Status", "Duration", "Timestamp", "Step Execution History"]
    for col_idx, text in enumerate(headers2, start=1):
        cell = ws2.cell(row=3, column=col_idx)
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_header

    # Populate Logs
    for row_idx, r in enumerate(results, start=4):
        ws2.cell(row=row_idx, column=1, value=r["test_id"]).alignment = Alignment(horizontal="center", vertical="top")
        ws2.cell(row=row_idx, column=2, value=r["name"]).alignment = Alignment(horizontal="left", vertical="top")
        ws2.cell(row=row_idx, column=3, value=r["description"]).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        status_cell = ws2.cell(row=row_idx, column=4, value=r["status"])
        status_cell.alignment = Alignment(horizontal="center", vertical="top")
        if r["status"] == "PASS":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        else:
            status_cell.fill = fail_fill
            status_cell.font = fail_font

        ws2.cell(row=row_idx, column=5, value=f"{r['duration_sec']}s").alignment = Alignment(horizontal="right", vertical="top")
        ws2.cell(row=row_idx, column=6, value=r["timestamp"]).alignment = Alignment(horizontal="center", vertical="top")
        
        # Large text wrap for logs
        log_cell = ws2.cell(row=row_idx, column=7, value=r["log"])
        log_cell.alignment = wrap_align

        # Style layout borders
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, 8):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.border = border_all
            cell.font = regular_font
            if is_even and col_idx != 4:  # Avoid overwriting status cell color
                cell.fill = zebra_fill

    # Set defined column widths for the logs sheet
    ws2.column_dimensions["A"].width = 10   # ID
    ws2.column_dimensions["B"].width = 25   # Name
    ws2.column_dimensions["C"].width = 30   # Description
    ws2.column_dimensions["D"].width = 12   # Status
    ws2.column_dimensions["E"].width = 12   # Duration
    ws2.column_dimensions["F"].width = 20   # Timestamp
    ws2.column_dimensions["G"].width = 80   # Execution History steps

    # Save Excel file
    wb.save(filepath)
    print(f"[SUCCESS] Excel report saved to: {os.path.abspath(filepath)}")

def generate_csv_fallback(results, filepath):
    """Fallback generator to save reports as a CSV if openpyxl isn't available."""
    import csv
    try:
        with open(filepath, mode="w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            # Write Header
            writer.writerow(["Test ID", "Test Name", "Description", "Status", "Duration (s)", "Timestamp", "Steps / Execution Log"])
            # Write rows
            for r in results:
                writer.writerow([
                    r["test_id"],
                    r["name"],
                    r["description"],
                    r["status"],
                    r["duration_sec"],
                    r["timestamp"],
                    r["log"].replace("\n", " | ")
                ])
        print(f"[SUCCESS] Fallback CSV report saved to: {os.path.abspath(filepath)}")
    except Exception as e:
        print(f"[ERROR] Failed to write fallback CSV report: {e}")
