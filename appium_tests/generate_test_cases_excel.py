#!/usr/bin/env python3
"""
Generate Excel file with all 400 test cases from the appium_tests directory.
"""

import sys
import os

# Add the appium_tests directory to the path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from categories.functional.test_functional import TEST_CASES as functional_cases
from categories.ui_ux.test_ui_ux import TEST_CASES as ui_ux_cases
from categories.compatibility.test_compatibility import TEST_CASES as compatibility_cases
from categories.performance.test_performance import TEST_CASES as performance_cases
from categories.security.test_security import TEST_CASES as security_cases
from categories.api.test_api import TEST_CASES as api_cases
from categories.database.test_database import TEST_CASES as database_cases
from categories.accessibility.test_accessibility import TEST_CASES as accessibility_cases
from categories.mobile_specific.test_mobile_specific import TEST_CASES as mobile_specific_cases
from categories.regression.test_regression import TEST_CASES as regression_cases
from categories.e2e.test_e2e import TEST_CASES as e2e_cases

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

# Combine all test cases
ALL_TEST_CASES = (
    functional_cases +
    ui_ux_cases +
    compatibility_cases +
    performance_cases +
    security_cases +
    api_cases +
    database_cases +
    accessibility_cases +
    mobile_specific_cases +
    regression_cases +
    e2e_cases
)

print(f"Total test cases: {len(ALL_TEST_CASES)}")

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Test Cases"

# Define headers
headers = ["Test ID", "Category", "Test Name", "Description", "Steps"]

# Apply header styling
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Write test cases
for row_num, test_case in enumerate(ALL_TEST_CASES, 2):
    # Test ID
    ws.cell(row=row_num, column=1, value=test_case.get("id", ""))
    
    # Category
    ws.cell(row=row_num, column=2, value=test_case.get("category", ""))
    
    # Test Name
    ws.cell(row=row_num, column=3, value=test_case.get("name", ""))
    
    # Description
    ws.cell(row=row_num, column=4, value=test_case.get("description", ""))
    
    # Steps - format as a readable string
    steps = test_case.get("steps", [])
    steps_text = "\n".join([f"{i+1}. {step[0]}: {step[1]}" if len(step) > 1 else f"{i+1}. {step[0]}" for i, step in enumerate(steps)])
    ws.cell(row=row_num, column=5, value=steps_text)

# Adjust column widths
ws.column_dimensions['A'].width = 15  # Test ID
ws.column_dimensions['B'].width = 25  # Category
ws.column_dimensions['C'].width = 40  # Test Name
ws.column_dimensions['D'].width = 50  # Description
ws.column_dimensions['E'].width = 60  # Steps

# Apply alignment to all data cells
for row in ws.iter_rows(min_row=2, max_row=len(ALL_TEST_CASES)+1):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# Freeze header row
ws.freeze_panes = 'A2'

# Save the Excel file
output_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_cases_400.xlsx")
wb.save(output_file)

print(f"Excel file generated successfully: {output_file}")
print(f"Total test cases exported: {len(ALL_TEST_CASES)}")
